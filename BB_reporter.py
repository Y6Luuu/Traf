import tkinter as tk
from os import path
from glob import glob  
import pandas as pd
from openpyxl import load_workbook

pd.reset_option('display.float_format')
pd.options.display.float_format = '{:.2f}'.format
#window creating
window = tk.Tk() 
window.title('BB reporter')
window.geometry('400x150')  # 这里的乘是小x

var_path = tk.StringVar()
entry_path = tk.Entry(window, textvariable=var_path, font=('Arial', 14))
entry_path.place(x=100,y=90)

# var_path2 = tk.StringVar()
# entry_path2 = tk.Entry(window, textvariable=var_path2, font=('Arial', 14))
# entry_path2.place(x=100,y=50)


def reporting():
    path = var_path.get()
    # path2 = var_path2.get()


    #database
    df_ffca= pd.read_excel(path + '/FFCA.xlsx' )
    df_qv = pd.read_excel(path + '/QV.xlsx')
    df_stock = pd.read_excel(path + '/stock.xlsx') 
    df_stock.drop([0])
    df_inventory = pd.read_excel(path + '/inventory.xlsx')
    df_WC = pd.read_excel(path + '/fixed_doc/Warehouse_Risk_Category.xlsx')
    df_CS = pd.read_excel(path + '/CS.xlsx')
    df_CP = pd.read_excel(path + '/fixed_doc/CP_approved.xlsx')
    df_movement = pd.read_excel(path + '/movement.xlsx')
    df_vessel = pd.read_excel(path + '/inventory2.xlsx')
    df_invoice = pd.read_excel(path + '/invoice.xlsx')
    df_input= pd.read_excel(path + '/input.xlsx')
    input_unpaid = pd.read_excel(path + '\input.xlsx', sheet_name = 'Sheet2')

    cols_ffca_1 = ['Purchase Assignment ID', 'status']
    cols_input = ['Purchase Assignment Reference', 'status']

    df_ffca_filter = pd.DataFrame(df_ffca, columns=cols_ffca_1)
    df_input = pd.DataFrame(df_input, columns=cols_input)

    # df_input = pd.merge(df_ffca_filter, df_input, left_on=["Purchase Assignment ID"], right_on=["assignment"], how= 'inner')
    # df_input.columns = ['Purchase Assignment Reference', 'status']

    # def expected(a):
    #     return a == "Expected"
    # def instorage(a):
    #     return a == "In-Storage"
    # def intransit(a):
    #     return a == "In-Transit"
    # def delivered(a):
    #     return a == "Delivered"

    # df1 = df_input.loc[df_input['status'].apply(expected)]
    # df2 = df_input.loc[df_input['status'].apply(instorage)]
    # df3 = df_input.loc[df_input['status'].apply(intransit)]
    # df4 = df_input.loc[df_input['status'].apply(delivered)]

    def prepayment(a):
        return a == "prepayment"
    def storage(a):
        return a == "storage"
    def transit(a):
        return a == "transit"
    def sales(a):
        return a == "sales"


    df1 = df_input.loc[df_input['status'].apply(prepayment)]
    df2 = df_input.loc[df_input['status'].apply(storage)]
    df3 = df_input.loc[df_input['status'].apply(transit)]
    df4 = df_input.loc[df_input['status'].apply(sales)]




    #prepayment
    cols_ffca = ['Purchase Assignment ID', 'Commodity', 'Inventory ID','Counterparty of Purchase', \
        'Purchase Settled Amount', 'Authorised Paid Quantity', 'Financed Quantity', 'Current Value', 'Treasury Payment Date', 'Purchase Invoice', 'Purchase Invoiced Amount']
    cols_qv = ['Assignment ID', 'Group Company', 'Trade ID', 'Quota ID', 'Traffic Operator']
    cols_inventory = ['INVENTORY ID', 'PURCHASE ASSIGNMENT', 'DRY WEIGHT', 'WET WEIGHT', 'EXPECTED RECEIPT DATE']
    cols_stock = ['Inventory Id','Origin Country']

    df_ffca_f1 = pd.DataFrame(df_ffca, columns=cols_ffca)
    df_qv_f1 = pd.DataFrame(df_qv, columns=cols_qv)
    df_inventory_f1 = pd.DataFrame(df_inventory, columns=cols_inventory)
    df_stock_f1 = pd.DataFrame(df_stock, columns=cols_stock)

    ffca_qv = pd.merge(df_ffca_f1, df_qv_f1, left_on=["Purchase Assignment ID"], right_on=["Assignment ID"], how= 'outer')
    inventory_stock = pd.merge(df_inventory_f1, df_stock_f1, left_on=['INVENTORY ID'], right_on=['Inventory Id'])
    final_prepayment = pd.merge(ffca_qv, inventory_stock, left_on=['Inventory ID'], right_on=['INVENTORY ID'])

    final_prepayment.columns = ['Purchase Assignment Reference', 'Purchase Commodity', 'Purchase Inventory ID', 'Purchase Contractual Counterparty', \
        'Purchase Settlement Valuation(USD)',  'Authorised Paid Quantity', 'Financed Quantity','Inventory Current Value(USD)',\
            'Purchase First Paid Date', 'Purchase Invoice ID', 'Purchase Invoice Amount(USD)','Assignment ID', 'Purchase Group Company Code', 'Purchase Trade ID', \
                'Purchase Quota Titan ID', 'Purchase Traffic Operator', 'INVENTORY ID', \
                        'PURCHASE ASSIGNMENT', 'Purchase Dry Weight', 'Purchase Wet Weight', 'Purchase Expected Receipt Date', 'Inventory Id', 'Purchase Material Origin Country']

    #增加列
    final_prepayment['Purchase Covered Amount(USD)'] = '0'
    final_prepayment['Purchase Dry Weight UOM'] = 'MT'
    final_prepayment['Purchase Wet Weight UOM'] = 'MT'
    final_prepayment['Receipt In Progress'] = 'N'
    final_prepayment['Delivery In Progress'] = 'N'
    final_prepayment['Authorised Paid Quantity'] = pd.to_numeric(final_prepayment['Authorised Paid Quantity'], errors='coerce')
    final_prepayment['Financed Quantity'] = pd.to_numeric(final_prepayment['Financed Quantity'], errors='coerce')

    final_prepayment.loc[:,"Financing Ratio"] = final_prepayment["Authorised Paid Quantity"]/final_prepayment["Financed Quantity"]

    final_prepayment.loc[:, 'Purchase Paid Amount(USD)'] = - final_prepayment['Purchase Invoice Amount(USD)']
    final_prepayment.loc[:, 'Purchase Uncovered Amount(USD)'] = final_prepayment['Purchase Invoice Amount(USD)']
    final_prepayment['Purchase Invoice Amount(USD)'] = - final_prepayment['Purchase Invoice Amount(USD)']

    #复制列
    final_prepayment.loc[:, 'Eligible Dry Weight'] = final_prepayment['Purchase Dry Weight']
    final_prepayment.loc[:, 'Eligible Wet Weight'] = final_prepayment['Purchase Wet Weight']
   

    #删除列
    final_prepayment.drop(columns=[ 'Assignment ID', 'INVENTORY ID', 'PURCHASE ASSIGNMENT', 'Inventory Id'])
    #修改顺序
    order = ['Purchase Group Company Code', 'Purchase Trade ID', 'Purchase Quota Titan ID', 'Purchase Assignment Reference', 'Purchase Inventory ID', \
        'Purchase Traffic Operator', 'Purchase Contractual Counterparty',  \
            'Purchase Commodity', \
                'Purchase Dry Weight', 'Purchase Dry Weight UOM', 'Purchase Wet Weight', 'Purchase Wet Weight UOM', 'Purchase Material Origin Country', \
                    'Purchase Settlement Valuation(USD)', 'Purchase First Paid Date', 'Purchase Expected Receipt Date', 'Purchase Invoice ID', \
                        'Purchase Invoice Amount(USD)', 'Purchase Paid Amount(USD)', 'Authorised Paid Quantity', 'Financed Quantity', 'Financing Ratio', \
                            'Purchase Covered Amount(USD)', 'Purchase Uncovered Amount(USD)', 'Receipt In Progress', 'Delivery In Progress', 'Eligible Dry Weight', 'Eligible Wet Weight']

    final_prepayment = final_prepayment[order]

    result1 = pd.merge(final_prepayment, df1, on=["Purchase Assignment Reference"])
    # result1 = result1.drop_duplicates(['Purchase Assignment Reference'])

    import numpy as np
    test_p = result1.copy(deep=True)

    test_p['Purchase Contractual Counterparty'] = np.nan
    test_p['Purchase Settlement Valuation(USD)'] = np.nan
    test_p['Authorised Paid Quantity'] = np.nan
    test_p['Financed Quantity'] = np.nan
    test_p['Financing Ratio'] = np.nan
    test_p['Purchase Covered Amount(USD)'] = np.nan
    test_p['Purchase Uncovered Amount(USD)'] = np.nan
    test_p['Receipt In Progress'] = np.nan
    test_p['Delivery In Progress'] = np.nan
    test_p['Eligible Dry Weight'] = np.nan
    test_p['Eligible Wet Weight'] = np.nan
    test_p['Purchase Uncovered Amount(USD)'] = np.nan


    test_q = result1.copy(deep=True)
    test_q['Purchase Invoice ID'] = np.nan


    test_f1 = pd.concat([test_p,test_q])
    test_f1= test_f1.sort_values(by=['Purchase Assignment Reference'], ascending=False)

    test_f1[['Purchase Assignment Reference']] = test_f1[['Purchase Assignment Reference']].astype('str')
    test_f1['Purchase Trade ID'] = test_f1['Purchase Assignment Reference'].map(lambda x: str(x)[:6])
    test_f1.loc[:, 'Purchase Quota Titan ID'] = test_f1['Purchase Assignment Reference'].apply(lambda st: st[st.rfind(".")-8: st.rfind(".")])

    # test_f1.to_excel(path + "\output.xlsx", sheet_name='prepayment')
    writer = pd.ExcelWriter(path + '/final_test.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    test_f1.to_excel(excel_writer=writer, sheet_name='prepayment')
    writer.save()
    writer.close()


        #storage
    #column filter
    cols_ffca = ['Purchase Assignment ID', 'Commodity', 'Inventory ID','Counterparty of Purchase', \
        'Purchase Settled Amount', 'Collateral Doc Type', 'Authorised Paid Quantity', 'Financed Quantity', 'Current Value']
    cols_qv = ['Assignment ID', 'Group Company', 'Trade ID', 'Quota ID', 'Traffic Operator', 'Collateral Doc Date', 'Country']
    cols_inventory = ['INVENTORY ID','COLLATERAL DOCUMENT DATE', 'PURCHASE ASSIGNMENT', 'DRY WEIGHT', 'WET WEIGHT', 'CITY']
    cols_stock = ['Inventory Id','Origin Country']
    cols_warehouse = ['INVENTORY ID', 'LOCATION']
    cols_wc = ['LOCATION_DESCRIPTION', 'RISK_CATEGORY']

    cols_ffca_blend = ['Purchase Assignment ID', 'Commodity', 'Inventory ID','Counterparty of Purchase', \
        'Purchase Settled Amount', 'Collateral Doc Type', 'Authorised Paid Quantity', 'Financed Quantity', 'Current Value', 'QTY']

    #NN
    # cols_stock = ['Inventory Id','Origin Country', 'Inventory Quantity(Wet)', 'Delivery In Progress Qty', 'Receipt In Progress Qty (Wet)']

    # df_inventory['INVENTORY ID'] = df_inventory['INVENTORY ID'].astype(str)
    # df_stock['Inventory Id'] = df_stock['Inventory Id'].astype(str)
    # df_ffca['Inventory ID'] = df_ffca['Inventory ID'].astype(str)

    df_stock['Receipt In Progress Qty (Wet)'] = pd.to_numeric(df_stock['Receipt In Progress Qty (Wet)'], errors='coerce')
    df_stock['Receipt In Progress Qty (Wet)'] = df_stock['Receipt In Progress Qty (Wet)'].fillna(0)
    df_warehouse = pd.DataFrame(df_vessel, columns=cols_warehouse)


    ndf_ffca_f1 = pd.DataFrame(df_ffca, columns=cols_ffca)
    df_qv_f1 = pd.DataFrame(df_qv, columns=cols_qv)
    df_inventory_f1 = pd.DataFrame(df_inventory, columns=cols_inventory)
    df_stock_f1 = pd.DataFrame(df_stock, columns=cols_stock)
    df_WC_f1 = pd.DataFrame(df_WC, columns=cols_wc)
    df_WC_f1['RISK_CATEGORY'] = df_WC_f1['RISK_CATEGORY'].map(lambda x: str(x)[13:])

    df_inventory_f1 = pd.merge(df_inventory_f1, df_warehouse, on='INVENTORY ID')


    #blend
    bdf_ffca_f1 = pd.DataFrame(df_ffca, columns=cols_ffca_blend)

    def blend(a):
            return a == "Blend"

    df_blend_inventory = df_inventory_f1.loc[df_inventory_f1['PURCHASE ASSIGNMENT'].apply(blend)]

    row_list = df_inventory_f1[df_inventory_f1['PURCHASE ASSIGNMENT'] == 'Blend'].index.tolist() # 获得含有该值的行的行号
    df_nblend_inventory = df_inventory_f1.drop(row_list)

    bffca_qv = pd.merge(bdf_ffca_f1, df_qv_f1, left_on=["Purchase Assignment ID"], right_on=["Assignment ID"], how= 'left')
    binventory_stock = pd.merge(df_blend_inventory, df_stock_f1, left_on=['INVENTORY ID'], right_on=['Inventory Id'], how='left')
    bfinal_BB_WHS = pd.merge(bffca_qv, binventory_stock, left_on=['Inventory ID'], right_on=['INVENTORY ID'])
    bfinal_BB_WHS = pd.merge(bfinal_BB_WHS,df_WC_f1, left_on = 'LOCATION', right_on= 'LOCATION_DESCRIPTION', how='outer')


    def function(a, b, c):
        return c * (1 + (a-b)/a)

    bfinal_BB_WHS['Purchased Wet Weight']  = bfinal_BB_WHS.apply(lambda x: function(x['WET WEIGHT'], x['DRY WEIGHT'], x['QTY']), axis = 1)

    bfinal_BB_WHS = bfinal_BB_WHS.drop(columns = ['WET WEIGHT', 'DRY WEIGHT'])

    #改名
    bfinal_BB_WHS.columns = ['Purchase Assignment Reference', 'Purchase Commodity', 'Purchase Inventory ID', 'Purchase Contractual Counterparty', \
        'Purchase Settlement Valuation(USD)', 'Collateral Document Type', 'Authorised Paid Quantity', \
            'Financed Quantity', 'Inventory Current Value(USD)', 'Purchase Dry Weight' , 'Assignment ID', 'Purchase Group Company Code', 'Purchase Trade ID', \
                'Purchase Quota Titan ID', 'Purchase Traffic Operator', 'Collateral Document Date',  'Warehouse Country', \
                        'INVENTORY ID',  'COLLATERAL DOCUMENT DATE', 'PURCHASE ASSIGNMENT', 'Warehouse City', 'Warehouse Reference', 'Inventory Id','Purchase Material Origin Country',\
                            'LOCATION_DESCRIPTION', 'Warehouse Category',  'Purchase Wet Weight']

    bfinal_BB_WHS['Collateral Reference'] = None
    bfinal_BB_WHS['Purchase Dry Weight UOM'] = 'MT'
    bfinal_BB_WHS['Purchase Wet Weight UOM'] = 'MT'
    bfinal_BB_WHS['Receipt In Progress'] = 'N'
    bfinal_BB_WHS['Delivery In Progress'] = 'N'
    # bfinal_BB_WHS['Receipt In Progress'] = bfinal_BB_WHS.apply(lambda x: function(x['Receipt In Progress Qty (Wet)'], x['Inventory Quantity(Wet)']), axis = 1)
    # bfinal_BB_WHS['Delivery In Progress'] = bfinal_BB_WHS.apply(lambda x: function(x['Delivery In Progress Qty'], x['Inventory Quantity(Wet)']), axis = 1)

    bfinal_BB_WHS['Authorised Paid Quantity'] = pd.to_numeric(bfinal_BB_WHS['Authorised Paid Quantity'], errors='coerce')
    bfinal_BB_WHS['Financed Quantity'] = pd.to_numeric(bfinal_BB_WHS['Financed Quantity'], errors='coerce')
    bfinal_BB_WHS.loc[:,"Financing Ratio"] = bfinal_BB_WHS["Authorised Paid Quantity"]/bfinal_BB_WHS["Financed Quantity"]

    # bfinal_BB_WHS[['Authorized Paid Quantity', 'Financed Quantity']] = bfinal_BB_WHS[['Authorized Paid Quantity', 'Financed Quantity']].astype('float')
    # bfinal_BB_WHS.loc[:,"Financing Ratio"] = bfinal_BB_WHS["Authorized Paid Quantity"]/bfinal_BB_WHS["Financed Quantity"]

    #复制列
    bfinal_BB_WHS.loc[:, 'Eligible Dry Weight'] = bfinal_BB_WHS['Purchase Dry Weight']
    bfinal_BB_WHS.loc[:, 'Eligible Wet Weight'] = bfinal_BB_WHS['Purchase Wet Weight']

    #删除列
    # bfinal_BB_WHS.drop(columns=['Inventory Quantity(Wet)', 'Delivery In Progress Qty', 'Receipt In Progress Qty (Wet)', 'LOCATION_DESCRIPTION'])
    bfinal_BB_WHS.drop(columns=[ 'LOCATION_DESCRIPTION'])
    #修改顺序
    order = ['Purchase Group Company Code', 'Purchase Trade ID', 'Purchase Quota Titan ID', 'Purchase Assignment Reference', 'Purchase Inventory ID', \
        'Purchase Traffic Operator', 'Purchase Contractual Counterparty', 'Collateral Document Type', 'Collateral Reference', 'COLLATERAL DOCUMENT DATE', \
            'Purchase Commodity', 'Purchase Material Origin Country', 'Warehouse Reference', 'Warehouse Category', 'Warehouse Country', 'Warehouse City', \
                'Purchase Dry Weight', \
                    'Purchase Dry Weight UOM', 'Purchase Wet Weight', 'Purchase Wet Weight UOM', 'Purchase Settlement Valuation(USD)',\
                        'Receipt In Progress', 'Delivery In Progress', 'Eligible Dry Weight', 'Eligible Wet Weight', 'Authorised Paid Quantity', \
                            'Financed Quantity','Financing Ratio', 'Inventory Current Value(USD)']

    bfinal_BB_WHS_test = bfinal_BB_WHS[order]



    #non-blend

    nffca_qv = pd.merge(ndf_ffca_f1, df_qv_f1, left_on=["Purchase Assignment ID"], right_on=["Assignment ID"], how= 'left')
    ninventory_stock = pd.merge(df_nblend_inventory, df_stock_f1, left_on=['INVENTORY ID'], right_on=['Inventory Id'])
    nfinal_BB_WHS = pd.merge(nffca_qv, ninventory_stock, left_on=['Inventory ID'], right_on=['INVENTORY ID'])
    nfinal_BB_WHS = pd.merge(nfinal_BB_WHS,df_WC_f1, left_on = 'LOCATION', right_on= 'LOCATION_DESCRIPTION', how='outer')

    #改名
    nfinal_BB_WHS.columns = ['Purchase Assignment Reference', 'Purchase Commodity', 'Purchase Inventory ID', 'Purchase Contractual Counterparty', \
        'Purchase Settlement Valuation(USD)', 'Collateral Document Type', 'Authorised Paid Quantity', \
            'Financed Quantity', 'Inventory Current Value(USD)', 'Assignment ID', 'Purchase Group Company Code', 'Purchase Trade ID', \
                'Purchase Quota Titan ID', 'Purchase Traffic Operator', 'Collateral Document Date', 'Warehouse Country', \
                        'INVENTORY ID',  'COLLATERAL DOCUMENT DATE', 'PURCHASE ASSIGNMENT', 'Purchase Dry Weight', 'Purchase Wet Weight', 'Warehouse City', 'Warehouse Reference', 'Inventory Id','Purchase Material Origin Country',\
                            'LOCATION_DESCRIPTION', 'Warehouse Category']


    # #新增固定列
    # def function(a,b):
    #     if a == b:
    #         return 'N'
    #     elif a == 0:
    #         return 'N'
    #     else:
    #         return 'Y'
            
    nfinal_BB_WHS['Collateral Reference'] = None
    nfinal_BB_WHS['Purchase Dry Weight UOM'] = 'MT'
    nfinal_BB_WHS['Purchase Wet Weight UOM'] = 'MT'
    nfinal_BB_WHS['Receipt In Progress'] = 'N'
    nfinal_BB_WHS['Delivery In Progress'] = 'N'
    # nfinal_BB_WHS['Receipt In Progress'] = nfinal_BB_WHS.apply(lambda x: function(x['Receipt In Progress Qty (Wet)'], x['Inventory Quantity(Wet)']), axis = 1)
    # nfinal_BB_WHS['Delivery In Progress'] = nfinal_BB_WHS.apply(lambda x: function(x['Delivery In Progress Qty'], x['Inventory Quantity(Wet)']), axis = 1)

    nfinal_BB_WHS['Authorised Paid Quantity'] = pd.to_numeric(nfinal_BB_WHS['Authorised Paid Quantity'], errors='coerce')
    nfinal_BB_WHS['Financed Quantity'] = pd.to_numeric(nfinal_BB_WHS['Financed Quantity'], errors='coerce')
    nfinal_BB_WHS.loc[:,"Financing Ratio"] = nfinal_BB_WHS["Authorised Paid Quantity"]/nfinal_BB_WHS["Financed Quantity"]



    #复制列
    nfinal_BB_WHS.loc[:, 'Eligible Dry Weight'] = nfinal_BB_WHS['Purchase Dry Weight']
    nfinal_BB_WHS.loc[:, 'Eligible Wet Weight'] = nfinal_BB_WHS['Purchase Wet Weight']

    #删除列
    # nfinal_BB_WHS.drop(columns=['Inventory Quantity(Wet)', 'Delivery In Progress Qty', 'Receipt In Progress Qty (Wet)', 'LOCATION_DESCRIPTION'])
    nfinal_BB_WHS.drop(columns=[ 'LOCATION_DESCRIPTION'])
    #修改顺序
    order = ['Purchase Group Company Code', 'Purchase Trade ID', 'Purchase Quota Titan ID', 'Purchase Assignment Reference', 'Purchase Inventory ID', \
        'Purchase Traffic Operator', 'Purchase Contractual Counterparty', 'Collateral Document Type', 'Collateral Reference', 'COLLATERAL DOCUMENT DATE', \
            'Purchase Commodity', 'Purchase Material Origin Country', 'Warehouse Reference', 'Warehouse Category', 'Warehouse Country', 'Warehouse City', \
                'Purchase Dry Weight', \
                    'Purchase Dry Weight UOM', 'Purchase Wet Weight', 'Purchase Wet Weight UOM', 'Purchase Settlement Valuation(USD)',\
                        'Receipt In Progress', 'Delivery In Progress', 'Eligible Dry Weight', 'Eligible Wet Weight', 'Authorised Paid Quantity', \
                            'Financed Quantity','Financing Ratio', 'Inventory Current Value(USD)']

    nfinal_BB_WHS_test = nfinal_BB_WHS[order]  

    final_BB_WHS = pd.concat([nfinal_BB_WHS_test,bfinal_BB_WHS_test])


    result2 = pd.merge(final_BB_WHS, df2, on=["Purchase Assignment Reference"], how='right')
    result2 = result2.drop_duplicates(['Purchase Assignment Reference'])
    result2['Purchase Group Company Code'] = 'PTE'
    result2[['Purchase Assignment Reference']] = result2[['Purchase Assignment Reference']].astype('str')
    result2['Purchase Trade ID'] = result2['Purchase Assignment Reference'].map(lambda x: str(x)[:6])
    result2.loc[:, 'Purchase Quota Titan ID'] = result2['Purchase Assignment Reference'].apply(lambda st: st[st.rfind(".")-8: st.rfind(".")])

    


    # result2.to_excel(path + "\output.xlsx", sheet_name='storage')
    writer = pd.ExcelWriter(path + '/final_test.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    result2.to_excel(excel_writer=writer, sheet_name='storage')
    writer.save()
    writer.close()

        #transit
    #column filter
    cols_ffca = ['Purchase Assignment ID', 'Commodity', 'Inventory ID','Counterparty of Purchase', 'QTY', \
        'Purchase Settled Amount', 'Collateral Doc Type', 'Financing Bank facility', 'Purchase Settled Amount', 'Authorised Paid Quantity', 'Financed Quantity', 'Current Value']

    cols_qv = ['Assignment ID', 'Group Company', 'Trade ID', 'Quota ID', 'Traffic Operator', 'Collateral Doc Date', 'Transport Mode']

    cols_movement = ['ID', 'LOAD DATE','LOAD', 'DESTINATION', 'DESTINATION DATE']

    cols_inventory = ['INVENTORY ID', 'COLLATERAL DOCUMENT DATE', 'PURCHASE ASSIGNMENT', 'MOVEMENT ID','RECEIPT INCOTERMS', 'DRY WEIGHT', 'WET WEIGHT']

    cols_stock = ['Inventory Id','Origin Country']

    cols_vessel = ['TRANSPORT', 'INVENTORY ID']


    #转换数据类型

    df_stock_f1 = pd.DataFrame(df_stock, columns=cols_stock)
    df_vessel_f1 = pd.DataFrame(df_vessel, columns=cols_vessel)
    df_ffca_f1 = pd.DataFrame(df_ffca, columns=cols_ffca)
    df_qv_f1 = pd.DataFrame(df_qv, columns=cols_qv)
    df_movement_f1 = pd.DataFrame(df_movement, columns=cols_movement)
    df_inventory_f1 = pd.DataFrame(df_inventory, columns=cols_inventory)
    df_movement_f1['LOAD'] = df_movement_f1['LOAD'].astype(str)

    #deliver detail
    df_movement_f1.loc[:, 'Load Country'] = df_movement_f1['LOAD'].apply(lambda st: st[st.rfind("(")+1:st.rfind(")")])
    df_movement_f1.loc[:, 'Delivery Country'] = df_movement_f1['DESTINATION'].apply(lambda st: st[st.rfind("(")+1:st.find(")")])

    def blend(a):
        return a == "Blend"

    df_blend_inventory = df_inventory_f1.loc[df_inventory_f1['PURCHASE ASSIGNMENT'].apply(blend)]

    row_list = df_inventory_f1[df_inventory_f1['PURCHASE ASSIGNMENT'] == 'Blend'].index.tolist() # 获得含有该值的行的行号
    df_nblend_inventory = df_inventory_f1.drop(row_list)

    ffca_qv = pd.merge(df_ffca_f1, df_qv_f1, left_on=["Purchase Assignment ID"], right_on=["Assignment ID"], how= 'left')
    bmovement_inventory = pd.merge(df_movement_f1, df_blend_inventory, left_on=["ID"], right_on=["MOVEMENT ID"])
    bMIS = pd.merge(bmovement_inventory, df_stock_f1, left_on=['INVENTORY ID'], right_on=['Inventory Id'])
    bfinal_BB1 = pd.merge(ffca_qv, bMIS, left_on=['Inventory ID'], right_on=['INVENTORY ID'])
    bfinal_BB = pd.merge(bfinal_BB1, df_vessel_f1, on=['INVENTORY ID'])


    def function(a, b, c):
        return c * (1 + (a-b)/a)

    bfinal_BB['Purchase Wet Weight']  = bfinal_BB.apply(lambda x: function(x['WET WEIGHT'], x['DRY WEIGHT'], x['QTY']), axis = 1)

    bfinal_BB = bfinal_BB.drop(columns = ['WET WEIGHT', 'DRY WEIGHT'])



    #改名
    bfinal_BB.columns = ['Purchase Assignment Reference', 'Purchase Commodity', 'Purchase Inventory ID', 'Purchase Contractual Counterparty', \
        'Purchase Dry Weight', 'Purchase Settlement Valuation(USD)', 'Collateral Document Type', 'Financing Bank facility', 'Purchase Settled Amount', 'Authorised Paid Quantity', \
            'Financed Quantity', 'In Transit Current Value(USD)', 'Assignment ID', 'Purchase Group Company Code', 'Purchase Trade ID', \
                'Purchase Quota Titan ID', 'Purchase Traffic Operator', 'Collateral Document Date', 'Purchase Transport Mode', 'ID', \
                    'Load Date', 'Load Location', 'Delivery Location', 'Expected Delivery Date', 'Load Country', 'Destination Country', \
                        'INVENTORY ID', 'COLLATERAL DOCUMENT DATE', 'PURCHASE ASSIGNMENT', 'MOVEMENT ID', 'Purchase Incoterm', 'Inventory Id','Purchase Material Origin Country','Vessel Name', 'Purchase Wet Weight']


    #新增固定列
    bfinal_BB['Collateral Reference'] = None
    bfinal_BB['Purchase Dry Weight UOM'] = 'MT'
    bfinal_BB['Purchase Wet Weight UOM'] = 'MT'
    bfinal_BB['Receipt In Progress'] = 'N'
    bfinal_BB['Delivery In Progress'] = 'N'
    # bfinal_BB['Receipt In Progress'] = bfinal_BB.apply(lambda x: function(x['Receipt In Progress Qty (Wet)'], x['Inventory Quantity(Wet)']), axis = 1)
    # bfinal_BB['Delivery In Progress'] = bfinal_BB.apply(lambda x: function(x['Delivery In Progress Qty'], x['Inventory Quantity(Wet)']), axis = 1)
    # bfinal_BB[['Authorized Paid Quantity', 'Financed Quantity']] = bfinal_BB[['Authorized Paid Quantity', 'Financed Quantity']].astype('float')
    # bfinal_BB.loc[:,"Financing Ratio"] = bfinal_BB["Authorized Paid Quantity"]/bfinal_BB["Financed Quantity"]

    bfinal_BB['Authorised Paid Quantity'] = pd.to_numeric(bfinal_BB['Authorised Paid Quantity'], errors='coerce')
    bfinal_BB['Financed Quantity'] = pd.to_numeric(bfinal_BB['Financed Quantity'], errors='coerce')

    bfinal_BB.loc[:,"Financing Ratio"] = bfinal_BB["Authorised Paid Quantity"]/bfinal_BB["Financed Quantity"]

    #复制列
    bfinal_BB.loc[:, 'Eligible Dry Weight'] = bfinal_BB['Purchase Dry Weight']
    bfinal_BB.loc[:, 'Eligible Wet Weight'] = bfinal_BB['Purchase Wet Weight']
    bfinal_BB.loc[:, 'BL Date'] = bfinal_BB['COLLATERAL DOCUMENT DATE']


    #删除不要的列
    bfinal_BB.drop(columns=['Financing Bank facility', 'Purchase Settled Amount', 'Assignment ID', 'ID', \
        'INVENTORY ID', 'PURCHASE ASSIGNMENT', 'Collateral Document Date', 'MOVEMENT ID'])

    #修改顺序
    order = ['Purchase Group Company Code', 'Purchase Trade ID', 'Purchase Quota Titan ID', 'Purchase Assignment Reference', 'Purchase Inventory ID', \
        'Purchase Traffic Operator', 'Purchase Contractual Counterparty', 'Collateral Document Type', 'Collateral Reference', 'COLLATERAL DOCUMENT DATE', \
            'Purchase Commodity', 'Purchase Material Origin Country', 'Purchase Incoterm', 'Purchase Transport Mode', 'Vessel Name', 'Load Date', \
                'Load Location', 'Load Country', 'BL Date', 'Expected Delivery Date', 'Delivery Location', 'Destination Country', 'Purchase Dry Weight', \
                    'Purchase Dry Weight UOM', 'Purchase Wet Weight', 'Purchase Wet Weight UOM', 'Purchase Settlement Valuation(USD)',\
                        'Receipt In Progress', 'Delivery In Progress', 'Eligible Dry Weight', 'Eligible Wet Weight', 'Authorised Paid Quantity', \
                            'Financed Quantity','Financing Ratio', 'In Transit Current Value(USD)']
    bfinal_BB_test = bfinal_BB[order]

    movement_inventory = pd.merge(df_movement_f1, df_inventory_f1, left_on=["ID"], right_on=["MOVEMENT ID"], how='outer')
    MIS = pd.merge(movement_inventory, df_stock_f1, left_on=['INVENTORY ID'], right_on=['Inventory Id'])
    final_BB1 = pd.merge(ffca_qv, MIS, left_on=['Inventory ID'], right_on=['INVENTORY ID'])
    final_BB = pd.merge(final_BB1, df_vessel_f1, on=['INVENTORY ID'])

    # #新增固定列
    # def function(a,b):
    #     if a == b:
    #         return 'N'
    #     elif a == 0:
    #         return 'N'
    #     else:
    #         return 'Y'

    #改名
    final_BB.columns = ['Purchase Assignment Reference', 'Purchase Commodity', 'Purchase Inventory ID', 'Purchase Contractual Counterparty', \
        'QTY', 'Purchase Settlement Valuation(USD)', 'Collateral Document Type', 'Financing Bank facility', 'Purchase Settled Amount', 'Authorised Paid Quantity', \
            'Financed Quantity', 'In Transit Current Value(USD)', 'Assignment ID', 'Purchase Group Company Code', 'Purchase Trade ID', \
                'Purchase Quota Titan ID', 'Purchase Traffic Operator', 'Collateral Document Date', 'Purchase Transport Mode', 'ID', \
                    'Load Date', 'Load Location', 'Delivery Location', 'Expected Delivery Date', 'Load Country', 'Destination Country', \
                        'INVENTORY ID', 'COLLATERAL DOCUMENT DATE', 'PURCHASE ASSIGNMENT', 'MOVEMENT ID', 'Purchase Incoterm', 'Purchase Dry Weight', 'Purchase Wet Weight','Inventory Id','Purchase Material Origin Country','Vessel Name']


    #新增固定列
    final_BB['Collateral Reference'] = None
    final_BB['Purchase Dry Weight UOM'] = 'MT'
    final_BB['Purchase Wet Weight UOM'] = 'MT'
    final_BB['Receipt In Progress'] = 'N'
    final_BB['Delivery In Progress'] = 'N'
    # final_BB['Receipt In Progress'] = final_BB.apply(lambda x: function(x['Receipt In Progress Qty (Wet)'], x['Inventory Quantity(Wet)']), axis = 1)
    # final_BB['Delivery In Progress'] = final_BB.apply(lambda x: function(x['Delivery In Progress Qty'], x['Inventory Quantity(Wet)']), axis = 1)
    # final_BB[['Authorized Paid Quantity', 'Financed Quantity']] = final_BB[['Authorized Paid Quantity', 'Financed Quantity']].astype('float')
    # final_BB.loc[:,"Financing Ratio"] = final_BB["Authorized Paid Quantity"]/final_BB["Financed Quantity"]

    final_BB['Authorised Paid Quantity'] = pd.to_numeric(final_BB['Authorised Paid Quantity'], errors='coerce')
    final_BB['Financed Quantity'] = pd.to_numeric(final_BB['Financed Quantity'], errors='coerce')

    final_BB.loc[:,"Financing Ratio"] = final_BB["Authorised Paid Quantity"]/final_BB["Financed Quantity"]

    #复制列
    final_BB.loc[:, 'Eligible Dry Weight'] = final_BB['Purchase Dry Weight']
    final_BB.loc[:, 'Eligible Wet Weight'] = final_BB['Purchase Wet Weight']
    final_BB.loc[:, 'BL Date'] = final_BB['COLLATERAL DOCUMENT DATE']


    #删除不要的列
    final_BB.drop(columns=['QTY', 'Financing Bank facility', 'Purchase Settled Amount', 'Assignment ID', 'ID', \
        'INVENTORY ID', 'PURCHASE ASSIGNMENT', 'Collateral Document Date', 'MOVEMENT ID'])

    #修改顺序
    order = ['Purchase Group Company Code', 'Purchase Trade ID', 'Purchase Quota Titan ID', 'Purchase Assignment Reference', 'Purchase Inventory ID', \
        'Purchase Traffic Operator', 'Purchase Contractual Counterparty', 'Collateral Document Type', 'Collateral Reference', 'COLLATERAL DOCUMENT DATE', \
            'Purchase Commodity', 'Purchase Material Origin Country', 'Purchase Incoterm', 'Purchase Transport Mode', 'Vessel Name', 'Load Date', \
                'Load Location', 'Load Country', 'BL Date', 'Expected Delivery Date', 'Delivery Location', 'Destination Country', 'Purchase Dry Weight', \
                    'Purchase Dry Weight UOM', 'Purchase Wet Weight', 'Purchase Wet Weight UOM', 'Purchase Settlement Valuation(USD)',\
                        'Receipt In Progress', 'Delivery In Progress', 'Eligible Dry Weight', 'Eligible Wet Weight', 'Authorised Paid Quantity', \
                            'Financed Quantity','Financing Ratio', 'In Transit Current Value(USD)']
    final_BB_test = final_BB[order]
    final_BB_test

    final_BB = pd.concat([bfinal_BB_test,final_BB_test])

    result3 = pd.merge(final_BB, df3, on=["Purchase Assignment Reference"], how='right')
    result3 = result3.drop_duplicates(['Purchase Assignment Reference'])
    result3[['Purchase Assignment Reference']] = result3[['Purchase Assignment Reference']].astype('str')
    result3['Purchase Trade ID'] = result3['Purchase Assignment Reference'].map(lambda x: str(x)[:6])
    result3.loc[:, 'Purchase Quota Titan ID'] = result3['Purchase Assignment Reference'].apply(lambda st: st[st.rfind(".")-8: st.rfind(".")])

    # result3.to_excel(path + "\output.xlsx", sheet_name='transit')
    writer = pd.ExcelWriter(path + '/final_test.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    result3.to_excel(excel_writer=writer, sheet_name='transit')
    writer.save()
    writer.close()

    #sales




    df_pref = pd.merge(df4, df_ffca, left_on='Purchase Assignment Reference', right_on='Purchase Assignment ID', how='left')
    cols=['Purchase Assignment Reference', 'Sales Assignment ID ']
    df4_x = df4
    df_sref = pd.DataFrame(df_pref, columns=cols)

    dfnan = df_sref[df_sref['Sales Assignment ID '].isnull()]
    dfnonnan = df_sref[df_sref['Sales Assignment ID '].notnull()]

    df_candidates = pd.merge(dfnonnan, df_ffca, on='Sales Assignment ID ', how='left')
    df_candidates = pd.concat([df_candidates, dfnan])
    # # df_candidates = df_candidates[df_candidates['Financing Bank facility']== 'CONCS BB ING']
    df_candidates['status'] = 'sales'
    cols_candidate = ['Purchase Assignment Reference', 'status']

    df_candidates = pd.DataFrame(df_candidates, columns=cols_candidate)
    df_candidates = df_candidates.drop_duplicates(['Purchase Assignment Reference'])
    # df_candidates.columns = ['Purchase Assignment Reference', 'status']
    df4 = df_candidates



    #column filter
    cols_ffca = ['Purchase Assignment ID', 'Sales Assignment ID ', 'Commodity', 'Inventory ID','Counterparty of Purchase', 'Counterparty of Sales', 'Sales Invoice ID', 'Estimated Sales Valuation', \
    'Purchase Settled Amount', 'Authorised Paid Quantity', 'Financed Quantity', 'Current Value', 'Treasury Payment Date', 'Purchase Invoice', 'Purchase Invoiced Amount', \
        'Sales Invoiced Amount', 'Sales Settled Amount', 'QTY']
    cols_qv = ['Assignment ID', 'Group Company', 'Trade ID', 'Quota ID', 'Traffic Operator']
    cols_inventory = ['INVENTORY ID', 'PURCHASE ASSIGNMENT', 'DRY WEIGHT', 'WET WEIGHT', 'EXPECTED RECEIPT DATE']
    cols_stock = ['Inventory Id','Origin Country']
    cols_cs = ['ID', 'Assignments', 'Term', 'Issuing/Collecting Bank', '1st Confirming Bank', 'Opening Date', 'Expiry Date', 'Balance Amount', 'Maximum Amount']
    cols_ffca_blend = ['Purchase Assignment ID', 'Sales Assignment ID ', 'Commodity', 'Inventory ID','Counterparty of Purchase', 'Counterparty of Sales', 'Sales Invoice ID', 'Estimated Sales Valuation', \
    'Purchase Settled Amount', 'Authorised Paid Quantity', 'Financed Quantity', 'Current Value', 'Treasury Payment Date', 'Purchase Invoice', 'Purchase Invoiced Amount', \
        'Sales Invoiced Amount', 'Sales Settled Amount', 'QTY']
    cols_invoice = ['Assignment Ref', 'Invoice Type', 'Invoice Date', 'Payment Expected Due Date', 'Days Overdue', 'Invoice No', 'Unpaid Amount ', 'Invoice Amount']
    cols_ar = ['Sales Assignment ID ', 'Accounts Receivable']

    df_ffca_f1 = pd.DataFrame(df_ffca, columns=cols_ffca)
    df_qv_f1 = pd.DataFrame(df_qv, columns=cols_qv)
    df_inventory_f1 = pd.DataFrame(df_inventory, columns=cols_inventory)
    df_stock_f1 = pd.DataFrame(df_stock, columns=cols_stock)
    df_CS_f1 = pd.DataFrame(df_CS, columns=cols_cs)
    df_invoice_f1 = pd.DataFrame(df_invoice, columns=cols_invoice)
    df_ffca_f1['Current Value'] = pd.to_numeric(df_ffca_f1['Current Value'], errors='coerce')
    df_ffca_f1['Authorised Paid Quantity'] = pd.to_numeric(df_ffca_f1['Authorised Paid Quantity'], errors='coerce')
    df_ffca_f1['Financed Quantity'] = pd.to_numeric(df_ffca_f1['Financed Quantity'], errors='coerce'
                                                    )
    df_invoice_f1['Unpaid Amount '] = df_invoice_f1['Unpaid Amount '].str.replace(',', '').astype(str)
    df_invoice_f1['Unpaid Amount '] = pd.to_numeric(df_invoice_f1['Unpaid Amount '], errors='coerce')
    df_invoice_f1['Invoice Amount'] = df_invoice_f1['Invoice Amount'].str.replace(',', '').astype(str)
    df_invoice_f1['Invoice Amount'] = pd.to_numeric(df_invoice_f1['Invoice Amount'], errors='coerce')

    df_invoice_f1 = df_invoice_f1.loc[(df_invoice_f1['Invoice Type'] == 'FINAL') | (df_invoice_f1['Invoice Type'] == 'PROVISIONAL')]

    df_invoice_f1['Unpaid Amount '].replace('-', 0, inplace = True)
    df_invoice_f1['Paid Amount'] = df_invoice_f1['Invoice Amount'] - df_invoice_f1['Unpaid Amount ']
    df_ar = pd.DataFrame(df_ffca, columns=cols_ar)

    #sales
    #delete nan sales
    df_ffca_f1.dropna(subset=['Sales Assignment ID '], inplace=True)
    # 一、先将字段拆分
    df_CS_f1['Assignments']=df_CS['Assignments'].map(lambda x:x.split(','))
    #二、然后直接调用explode()方法
    df_CS_f1_explode = df_CS_f1.explode('Assignments')

    def blend(a):
        return a == "Blend"

    #blend
    bdf_ffca_f1 = pd.DataFrame(df_ffca, columns=cols_ffca_blend)
    df_blend_inventory = df_inventory_f1.loc[df_inventory_f1['PURCHASE ASSIGNMENT'].apply(blend)]

    row_list = df_inventory_f1[df_inventory_f1['PURCHASE ASSIGNMENT'] == 'Blend'].index.tolist() # 获得含有该值的行的行号
    df_inventory_f1 = df_inventory_f1.drop(row_list)

    bffca_qv = pd.merge(bdf_ffca_f1, df_qv_f1, left_on=["Purchase Assignment ID"], right_on=["Assignment ID"], how= 'left')
    bffca_qv_cp = pd.merge(bffca_qv, df_CP, left_on=['Counterparty of Sales'], right_on=['Sales Assignment Counterparty'], how='outer')
    binventory_stock = pd.merge(df_blend_inventory, df_stock_f1, left_on=['INVENTORY ID'], right_on=['Inventory Id'], how='left')
    FQC = pd.merge(bffca_qv_cp, df_CS_f1_explode, left_on=["Sales Assignment ID "], right_on=["Assignments"], how = 'outer')
    bfinal_sales = pd.merge(FQC, binventory_stock, left_on=['Inventory ID'], right_on=['INVENTORY ID'])

    def function(a, b, c):
        return c * (1 + (a-b)/a)

    bfinal_sales['Purchased Wet Weight']  = bfinal_sales.apply(lambda x: function(x['WET WEIGHT'], x['DRY WEIGHT'], x['QTY']), axis = 1)

    bfinal_sales = bfinal_sales.drop(columns = ['WET WEIGHT', 'DRY WEIGHT'])

    #non-blend
    ffca_qv = pd.merge(df_ffca_f1, df_qv_f1, left_on=["Purchase Assignment ID"], right_on=["Assignment ID"], how= 'left')
    ffca_qv_cp = pd.merge(ffca_qv, df_CP, left_on=['Counterparty of Sales'], right_on=['Sales Assignment Counterparty'], how='outer')
    inventory_stock = pd.merge(df_inventory_f1, df_stock_f1, left_on=['INVENTORY ID'], right_on=['Inventory Id'], how='left')
    nFQC = pd.merge(ffca_qv_cp, df_CS_f1_explode, left_on=["Sales Assignment ID "], right_on=["Assignments"], how = 'outer')
    nfinal_sales = pd.merge(nFQC, inventory_stock, left_on=['Inventory ID'], right_on=['INVENTORY ID'])
    nfinal_sales.columns = ['Purchase Assignment ID', 'Sales Assignment ID ', 'Commodity',
    'Inventory ID', 'Counterparty of Purchase', 'Counterparty of Sales',
    'Sales Invoice ID', 'Estimated Sales Valuation',
    'Purchase Settled Amount', 'Authorised Paid Quantity',
    'Financed Quantity', 'Current Value', 'Treasury Payment Date',
    'Purchase Invoice', 'Purchase Invoiced Amount', 'Sales Invoiced Amount',
    'Sales Settled Amount', 'QTY', 'Assignment ID', 'Group Company',
    'Trade ID', 'Quota ID', 'Traffic Operator',
    'Sales Assignment Counterparty',
    'Sales Contractual Counterparty Open Account Approved ', 'ID',
    'Assignments', 'Term', 'Issuing/Collecting Bank', '1st Confirming Bank',
    'Opening Date', 'Expiry Date', 'Balance Amount', 'Maximum Amount',
    'INVENTORY ID', 'PURCHASE ASSIGNMENT', 'Purchase Dry Weight', 'Purchase Wet Weight',
    'EXPECTED RECEIPT DATE', 'Inventory Id', 'Origin Country']

    order = ['Purchase Assignment ID', 'Sales Assignment ID ', 'Commodity','Inventory ID', 'Counterparty of Purchase', 'Counterparty of Sales', 'Sales Invoice ID', 'Estimated Sales Valuation', \
    'Purchase Settled Amount', 'Authorised Paid Quantity','Financed Quantity', 'Current Value', 'Treasury Payment Date','Purchase Invoice', 'Purchase Invoiced Amount', 'Sales Invoiced Amount','Sales Settled Amount',  \
        'Assignment ID', 'Group Company','Trade ID', 'Quota ID', 'Traffic Operator','Sales Assignment Counterparty','Sales Contractual Counterparty Open Account Approved ', 'ID', \
    'Assignments', 'Term', 'Issuing/Collecting Bank', '1st Confirming Bank','Opening Date', 'Expiry Date', 'Balance Amount', 'Maximum Amount', 'INVENTORY ID', 'PURCHASE ASSIGNMENT', 'QTY', 'Purchased Wet Weight',\
    'EXPECTED RECEIPT DATE', 'Inventory Id', 'Origin Country']

    bfinal_sales = bfinal_sales[order]

    bfinal_sales.columns = ['Purchase Assignment ID', 'Sales Assignment ID ', 'Commodity',
    'Inventory ID', 'Counterparty of Purchase', 'Counterparty of Sales',
    'Sales Invoice ID', 'Estimated Sales Valuation',
    'Purchase Settled Amount', 'Authorised Paid Quantity',
    'Financed Quantity', 'Current Value', 'Treasury Payment Date',
    'Purchase Invoice', 'Purchase Invoiced Amount', 'Sales Invoiced Amount',
    'Sales Settled Amount', 'Assignment ID', 'Group Company', 'Trade ID',
    'Quota ID', 'Traffic Operator', 'Sales Assignment Counterparty',
    'Sales Contractual Counterparty Open Account Approved ', 'ID',
    'Assignments', 'Term', 'Issuing/Collecting Bank', '1st Confirming Bank',
    'Opening Date', 'Expiry Date', 'Balance Amount', 'Maximum Amount',
    'INVENTORY ID', 'PURCHASE ASSIGNMENT', 'Purchase Dry Weight', 'Purchase Wet Weight',
    'EXPECTED RECEIPT DATE', 'Inventory Id', 'Origin Country']

    final_sales = pd.concat([nfinal_sales,bfinal_sales])

    final_sales = pd.merge(final_sales, df_invoice_f1, left_on=["Sales Assignment ID "], right_on=["Assignment Ref"], how='left')
    final_sales.fillna(0)


    # final_sales['Unpaid Amount ] = pd.to_numeric(final_sales['Unpaid Amount ], errors='coerce')
    final_sales['Estimated Sales Valuation'] = pd.to_numeric(final_sales['Estimated Sales Valuation'], errors='coerce')
    final_sales['Current Value'] = pd.to_numeric(final_sales['Current Value'], errors='coerce')

    #Comparison between invoice amount and current value


    def function1(a,b):
        if a == 'PROVISIONAL':
            return b
        else:
            return 0


    def function2(a,b):
        if a == 'FINAL':
            return b
        else:
            return 0

    def functionxx_3(a,b):
        if a > b:
            return a-b
        else:
            return 0


    final_sales['Unpaid Provisional Receivables(USD)'] = final_sales.apply(lambda x: function1(x['Invoice Type'], x['Unpaid Amount ']), axis = 1)
    final_sales['Unpaid Final Receivables(USD)'] = final_sales.apply(lambda x: function2(x['Invoice Type'], x['Unpaid Amount ']), axis = 1) #wrong
    # final_sales.loc[:,"Uninvoiced Final Receivables(USD)"] = final_sales["Estimated Sales Valuation"] - final_sales["Invoice Amount"]
    final_sales['Uninvoiced Final Receivables(USD)'] = final_sales.apply(lambda x: functionxx_3(x['Estimated Sales Valuation'], x['Invoice Amount']), axis = 1)


    def function3(a,b):
        if a >= b:
            return b
        if a < b:
            return -a

    def function4(a,b):
        if -a < b:
            return b + a
        if -a >= b:
            return 0

    def function5(a,b):
        if a >= b:
            return b
        else:
            return a

    def function6(a):
        if a == "FINAL":
            return "100%"
        else:
            return "90%"

    final_sales['Total Financeable Sales Invoices Covered by LC']  = final_sales.apply(lambda x: function3(x['Invoice Amount'], x['Balance Amount']), axis = 1)
    final_sales['LC Cover Available to SCR (SCR-Sales Contractual Rights)']  = final_sales.apply(lambda x: function4(x['Invoice Amount'], x['Balance Amount']), axis = 1)
    final_sales['Sales Provisional Percentage']  = final_sales.apply(lambda x: function6(x['Invoice Type']), axis = 1)
    final_sales['SCR Covered By LC (SCR-Sales Contractual Rights)']  = final_sales.apply(lambda x: function5(x['LC Cover Available to SCR (SCR-Sales Contractual Rights)'], x['Uninvoiced Final Receivables(USD)']), axis = 1)


    import numpy as np

    #adding 
    final_sales['Sales Provisional Number'] = '1'
    final_sales['Authorised Paid Quantity'] = pd.to_numeric(final_sales['Authorised Paid Quantity'], errors='coerce')
    final_sales['Financed Quantity'] = pd.to_numeric(final_sales['Financed Quantity'], errors='coerce')
    final_sales.loc[:,"Financing Ratio"] = final_sales["Authorised Paid Quantity"]/final_sales["Financed Quantity"]

    final_sales['Cost Code ID'] = 'I01'
    final_sales['Invoice Line Quantity'] = '0'
    final_sales['Invoice Line Quantity UOM'] = 'MT'
    final_sales['Purchase Dry Weight UOM'] = 'MT'
    final_sales['Purchase Wet Weight UOM'] = 'MT'
    final_sales['Receipt In Progress'] = 'N'
    final_sales['Delivery In Progress'] = 'N'
    final_sales['Summary Outstanding Receivables(USD)'] = final_sales['Unpaid Provisional Receivables(USD)'] + final_sales['Unpaid Final Receivables(USD)'] + final_sales['Uninvoiced Final Receivables(USD)']
    final_sales['Sales Invoice Counterparty'] = final_sales['Counterparty of Sales']
    final_sales['Eligible Dry Weight'] = final_sales['Purchase Dry Weight']
    final_sales['Eligible Wet Weight'] = final_sales['Purchase Wet Weight']


    def function7(a, b):
        if b == 0:
            return 0
        else:
            return -a/b
        
    def function8(a):
        if a>1:
            return 1
        else:
            return a
        


    final_sales['Sales Assignment Eligible Percentage']  = final_sales.apply(lambda x: function7(x['Estimated Sales Valuation'], x['Invoice Amount']), axis = 1)

    final_sales['Sales Assignment Eligible Percentage']  = final_sales.apply(lambda x: function8(x['Sales Assignment Eligible Percentage']), axis = 1)

    final_sales = final_sales.drop(columns=['Inventory ID', 'Current Value','Purchase Settled Amount', 'Purchase Invoice', 'Purchase Invoiced Amount', 'Current Value', 'Treasury Payment Date', \
    'Assignment ID', 'Sales Assignment Counterparty', 'Assignments', 'INVENTORY ID', 'PURCHASE ASSIGNMENT', 'Unpaid Amount ', \
    'EXPECTED RECEIPT DATE', 'Inventory Id', 'Origin Country', 'QTY', 'Assignment Ref', 'Sales Invoice ID', 'Sales Invoiced Amount', 'Sales Settled Amount'], axis=1)


    final_sales.columns = ['Purchase Assignment Reference', 'Sales Assignment Reference', 'Purchase Commodity',
    'Purchase Contractual Counterparty', 'Sales Assignment Counterparty', 
    'Cost Amount(USD)', 'Authorised Paid Quantity','Financed Quantity', 
    'Purchase Group Company Code', 'Purchase Trade ID', 'Purchase Quota Titan ID', 'Purchase Traffic Operator',
    'Sales Contractual Counterparty Open Account Approved', 'LC Reference', 'Sales Credit Security Type',
    'LC Opening Bank', 'LC Confirming Bank', 'LC Issuance Date', 'LC Expiry Date', 'LC Current Balance', 'LC Original Amount',
    'Purchase Dry Weight', 'Purchase Wet Weight', 
    'Sales Provisional Final', 'Sales Invoice Date', 'Sales Invoice Due Date', 'Sales Invoice Days Overdue','Sales Invoice ID','Invoice Line Amount(USD)','Paid Amount', 'Unpaid Provisional Receivables(USD)',
    'Unpaid Final Receivables(USD)', 'Uninvoiced Final Receivables(USD)',
    'Total Financeable Sales Invoices Covered by LC',
    'LC Cover Available to SCR (SCR-Sales Contractual Rights)',
    'Sales Provisional Percentage',
    'SCR Covered By LC (SCR-Sales Contractual Rights)',
    'Sales Provisional Number', 'Financing Ratio', 'Cost Code ID',
    'Invoice Line Quantity', 'Invoice Line Quantity UOM',
    'Purchase Dry Weight UOM', 'Purchase Wet Weight UOM',
    'Receipt In Progress', 'Delivery In Progress',
    'Summary Outstanding Receivables(USD)', 'Sales Invoice Counterparty',
    'Eligible Dry Weight', 'Eligible Wet Weight',
    'Sales Assignment Eligible Percentage']


    order4 = ['Purchase Group Company Code', 'Purchase Trade ID', 'Purchase Quota Titan ID', 'Purchase Assignment Reference', 'Sales Assignment Reference','Purchase Traffic Operator', \
    'Purchase Contractual Counterparty', 'Purchase Commodity', 'Sales Assignment Counterparty', 'Sales Credit Security Type', 'LC Reference', 'LC Original Amount', \
        'LC Current Balance', 'LC Issuance Date', 'LC Expiry Date', 'LC Opening Bank', 'LC Confirming Bank', 'Total Financeable Sales Invoices Covered by LC', \
                'LC Cover Available to SCR (SCR-Sales Contractual Rights)', 'SCR Covered By LC (SCR-Sales Contractual Rights)', 'Sales Invoice ID', 'Sales Provisional Final', \
                'Sales Provisional Number', 'Sales Provisional Percentage', 'Sales Invoice Date', 'Sales Invoice Due Date', 'Sales Invoice Days Overdue', 'Sales Invoice Counterparty', \
                    'Cost Code ID', 'Cost Amount(USD)', 'Invoice Line Quantity', 'Invoice Line Quantity UOM', 'Invoice Line Amount(USD)', 'Paid Amount', 'Sales Assignment Eligible Percentage', \
                            'Authorised Paid Quantity', 'Financed Quantity', 'Financing Ratio', 'Unpaid Provisional Receivables(USD)', 'Unpaid Final Receivables(USD)', 'Uninvoiced Final Receivables(USD)', \
                            'Summary Outstanding Receivables(USD)', 'Sales Contractual Counterparty Open Account Approved', 'Purchase Dry Weight', 'Purchase Dry Weight UOM', \
                                'Purchase Wet Weight', 'Purchase Wet Weight UOM', 'Receipt In Progress', 'Delivery In Progress', 'Eligible Dry Weight', 'Eligible Wet Weight']

    final_sales_test = final_sales[order4]


    result4 = pd.merge(final_sales_test, df4, on=["Purchase Assignment Reference"], how='right')
    result4['Cost Amount(USD)'] = pd.to_numeric(result4['Cost Amount(USD)'], errors='coerce')
    result4['Paid Amount'] = pd.to_numeric(result4['Paid Amount'], errors='coerce')
    result4 = result4.drop_duplicates(['Purchase Assignment Reference', 'Sales Invoice ID'])

    a = result4.groupby('Sales Invoice ID')['Purchase Dry Weight'].transform('sum')
    result4['%'] = (result4['Purchase Dry Weight']/result4.groupby('Sales Assignment Reference')['Purchase Dry Weight'].transform('sum'))

    result4['Sales Assignment Eligible Percentage'] = result4['%']

    result4['Unpaid Provisional Receivables(USD)'] = result4['Unpaid Provisional Receivables(USD)'] * result4['%']
    result4['Unpaid Final Receivables(USD)'] = result4['Unpaid Final Receivables(USD)'] * result4['%']
    result4['Uninvoiced Final Receivables(USD)'] = result4['Uninvoiced Final Receivables(USD)'] * result4['%']

    result4 = pd.merge(result4, df_ar, left_on='Sales Assignment Reference', right_on='Sales Assignment ID ')
    result4['Accounts Receivable'] = result4['Accounts Receivable'] * result4['%']
    result4['Summary Outstanding Receivables(USD)'] = result4['Accounts Receivable']
    result4 = result4.drop_duplicates(['Purchase Assignment Reference', 'Sales Invoice ID'])
    result4['Uninvoiced Final Receivables(USD)'] = result4['Summary Outstanding Receivables(USD)'] - result4['Unpaid Final Receivables(USD)']  - result4['Unpaid Provisional Receivables(USD)']

    result4 = result4.drop(columns = ['Sales Assignment ID ', 'Accounts Receivable'])

    dfa = result4.groupby(['Purchase Assignment Reference', 'Sales Assignment Reference'])['Invoice Line Amount(USD)'].sum()
    dfb = result4.groupby(['Purchase Assignment Reference', 'Sales Assignment Reference'])['Cost Amount(USD)'].sum()
    dfc = result4.groupby(['Purchase Assignment Reference', 'Sales Assignment Reference'])['Paid Amount'].sum()
    dfd = result4.groupby(['Purchase Assignment Reference', 'Sales Assignment Reference'])['Unpaid Provisional Receivables(USD)'].sum()
    dfe = result4.groupby(['Purchase Assignment Reference', 'Sales Assignment Reference'])['Unpaid Final Receivables(USD)'].sum()
    
    dfa.reset_index(drop=False)
    dfb.reset_index(drop=False)
    dfc.reset_index(drop=False)
    dfd.reset_index(drop=False)
    dfe.reset_index(drop=False)

    dfa = pd.DataFrame(dfa)
    dfb = pd.DataFrame(dfb)
    dfc = pd.DataFrame(dfc)
    dfd = pd.DataFrame(dfd)
    dfe = pd.DataFrame(dfe)

    df_1 = pd.merge(dfa,dfb, on = 'Purchase Assignment Reference')
    df_2 = pd.merge(dfd,dfc, on = 'Purchase Assignment Reference')
    df_3 = pd.merge(df_2,dfe, on = 'Purchase Assignment Reference')
    df_4 = pd.merge(df_2,df_3, on = 'Purchase Assignment Reference')
    df_5 = pd.merge(df_4,df_1, on = 'Purchase Assignment Reference')
    df_5 = df_5.drop_duplicates(['Unpaid Provisional Receivables(USD)_x'])

    df5 = df_5.reset_index()

    order = ['Purchase Group Company Code', 'Purchase Trade ID', 'Purchase Quota Titan ID', 'Purchase Assignment Reference', 'Sales Assignment Reference','Purchase Traffic Operator', \
    'Purchase Contractual Counterparty', 'Purchase Commodity', 'Sales Assignment Counterparty', 'Sales Credit Security Type', 'LC Reference', 'LC Original Amount', \
    'LC Current Balance', 'LC Issuance Date', 'LC Expiry Date', 'LC Opening Bank', 'LC Confirming Bank', 'Total Financeable Sales Invoices Covered by LC', \
        'LC Cover Available to SCR (SCR-Sales Contractual Rights)', 'SCR Covered By LC (SCR-Sales Contractual Rights)', 'Sales Invoice ID', 'Sales Provisional Final', \
                'Sales Provisional Number', 'Sales Provisional Percentage', 'Sales Invoice Date', 'Sales Invoice Due Date', 'Sales Invoice Days Overdue', 'Sales Invoice Counterparty', \
                'Cost Code ID', 'Cost Amount(USD)', 'Invoice Line Quantity', 'Invoice Line Quantity UOM', 'Invoice Line Amount(USD)', 'Paid Amount', 'Sales Assignment Eligible Percentage', \
                    'Authorised Paid Quantity', 'Financed Quantity', 'Financing Ratio', 'Unpaid Provisional Receivables(USD)', 'Unpaid Final Receivables(USD)', 'Uninvoiced Final Receivables(USD)', \
                            'Summary Outstanding Receivables(USD)', 'Sales Contractual Counterparty Open Account Approved', 'Purchase Dry Weight', 'Purchase Dry Weight UOM', \
                            'Purchase Wet Weight', 'Purchase Wet Weight UOM', 'Receipt In Progress', 'Delivery In Progress', 'Eligible Dry Weight', 'Eligible Wet Weight']
                            
    dff = pd.merge(df5, result4, on='Purchase Assignment Reference', how='right')
    dff1 = dff.drop(columns = ['Cost Amount(USD)_y', 'Paid Amount_y', 'Invoice Line Amount(USD)_y', 'Unpaid Final Receivables(USD)_y', 'Unpaid Provisional Receivables(USD)_y', 'Unpaid Provisional Receivables(USD)'])
    dff2 = dff1.drop_duplicates(['Purchase Assignment Reference'])
    dff3 = dff2.rename(columns={'Cost Amount(USD)_x':'Cost Amount(USD)', 'Paid Amount_x':'Paid Amount', 'Invoice Line Amount(USD)_x':'Invoice Line Amount(USD)', 'Unpaid Final Receivables(USD)_x':'Unpaid Final Receivables(USD)', 'Unpaid Provisional Receivables(USD)_x':'Unpaid Provisional Receivables(USD)'})
    dff = dff3 [order]





    dff['Sales Invoice ID'] = np.nan
    dff['Sales Provisional Final '] = np.nan
    dff['Sales Provisional Number'] = np.nan
    dff['Sales Provisional Percentage'] = np.nan
    dff['Sales Invoice Date'] = np.nan
    dff['Sales Invoice Due Date'] = np.nan
    dff['Sales Invoice Days Overdue'] = np.nan
    dff['Sales Invoice Counterparty'] = np.nan
    dff['Cost Code ID'] = np.nan
    dff['Invoice Line Quantity'] = np.nan
    dff['Invoice Line Quantity UOM'] = np.nan

    def function_x(a,b,c):
        if a == c:
            return 0
        if a > c and b > (a-c):
            return (a-c) 
        if a > c and b < a:
            return b

    def function_y(a,b,c):
        if a == c:
            return b
        if a > c and b > (a-c):
            return b - (a-c)
        if a > c and b < (a-c):
            return 0

    def function_z(a,b):
        if a >= b:
            return b
        else:
            return a

    dff = dff.loc[:, ~dff.columns.duplicated()]


    def function_x1(a,b,c):
        if a == 0 or b == 0:
            return c == 0
        else:
            return c

    def function_x2(a):
        if a == None:
            return a == 'Unknown'
        else:
            return a

    dff['Total Financeable Sales Invoices Covered by LC']  = dff.apply(lambda x: function_x(x['Invoice Line Amount(USD)'], x['LC Current Balance'], x['Paid Amount']), axis = 1)
    dff['LC Cover Available to SCR (SCR-Sales Contractual Rights)']  = dff.apply(lambda x: function_y(x['Invoice Line Amount(USD)'], x['LC Current Balance'], x['Paid Amount']), axis = 1)
    dff['Sales Provisional Percentage']  = dff.apply(lambda x: function6(x['Sales Provisional Final ']), axis = 1)
    dff['LC Cover Available to SCR (SCR-Sales Contractual Rights)'] = pd.to_numeric(dff['LC Cover Available to SCR (SCR-Sales Contractual Rights)'], errors='coerce') #chenge to float
    dff['SCR Covered By LC (SCR-Sales Contractual Rights)']  = dff.apply(lambda x: function5(x['LC Cover Available to SCR (SCR-Sales Contractual Rights)'], x['Uninvoiced Final Receivables(USD)']), axis = 1)

    dff["Financing Ratio"]  = dff.apply(lambda x: function_x1(x['Authorised Paid Quantity'], x['Financed Quantity'], x['Financing Ratio']), axis = 1)
    dff['LC Confirming Bank']  = dff.apply(lambda x: function_x2(x['LC Confirming Bank']), axis = 1)

    dff['Total Financeable Sales Invoices Covered by LC'].fillna(0, inplace=True)
    dff['LC Cover Available to SCR (SCR-Sales Contractual Rights)'].fillna(0, inplace=True)
    dff['SCR Covered By LC (SCR-Sales Contractual Rights)'].fillna(0, inplace=True)




    dff = dff.rename(columns={'Sales Provisional Final ':'Sales Provisional Final', })



    # dff['Summary Outstanding Receivables(USD)'] = dff['Unpaid Provisional Receivables(USD)'] + dff['Unpaid Final Receivables(USD)'] + dff['Uninvoiced Final Receivables(USD)']

    dff = dff[order4]
    dff.reset_index(drop=True)





    result4['LC Current Balance'] = np.nan
    result4['Sales Credit Security Type'] = np.nan
    result4['LC Reference'] = np.nan
    result4['LC Original Amount'] = np.nan
    result4['LC Issuance Date'] = np.nan
    result4['LC Expiry Date'] = np.nan
    result4['LC Confirming Bank'] = np.nan
    result4['Total Financeable Sales Invoices Covered by LC'] = np.nan
    result4['LC Cover Available to SCR (SCR-Sales Contractual Rights)'] = np.nan
    result4['SCR Covered By LC (SCR-Sales Contractual Rights)'] = np.nan
    result4['Cost Amount(USD)'] = np.nan
    result4['Authorised Paid Quantity'] = np.nan
    result4['Financed Quantity'] = np.nan
    result4['Uninvoiced Final Receivables(USD)'] = np.nan
    result4['Summary Outstanding Receivables(USD)'] = np.nan
    result4['Sales Contractual Counterparty Open Account Approved '] = np.nan
    result4['Purchase Dry Weight'] = np.nan
    result4['Purchase Dry Weight UOM'] = np.nan
    result4['Purchase Wet Weight'] = np.nan
    result4['Purchase Wet Weight UOM'] = np.nan
    result4['Receipt In Progress'] = np.nan
    result4['Delivery In Progress'] = np.nan
    result4['Eligible Dry Weight'] = np.nan
    result4['Eligible Wet Weight'] = np.nan

    result4 = result4.drop(columns = ['%', 'status'])
    result4 = result4[order4]
    result4.reset_index(drop=True)

    dff.columns = ['Purchase Group Company Code', 'Purchase Trade ID',
        'Purchase Quota Titan ID', 'Purchase Assignment Reference',
        'Sales Assignment Reference', 'Purchase Traffic Operator',
        'Purchase Contractual Counterparty', 'Purchase Commodity',
        'Sales Assignment Counterparty', 'Sales Credit Security Type',
        'LC Reference', 'LC Original Amount', 'LC Current Balance',
        'LC Issuance Date', 'LC Expiry Date', 'LC Opening Bank',
        'LC Confirming Bank', 'Total Financeable Sales Invoices Covered by LC',
        'LC Cover Available to SCR (SCR-Sales Contractual Rights)',
        'SCR Covered By LC (SCR-Sales Contractual Rights)', 'Sales Invoice ID',
        'Sales Provisional Final1', 'Sales Provisional Final2',
        'Sales Provisional Number', 'Sales Provisional Percentage',
        'Sales Invoice Date', 'Sales Invoice Due Date',
        'Sales Invoice Days Overdue', 'Sales Invoice Counterparty',
        'Cost Code ID', 'Cost Amount(USD)', 'Invoice Line Quantity',
        'Invoice Line Quantity UOM', 'Invoice Line Amount(USD)', 'Paid Amount',
        'Sales Assignment Eligible Percentage',
        'Authorised Paid Quantity', 'Financed Quantity', 'Financing Ratio',
        'Unpaid Provisional Receivables(USD)', 'Unpaid Final Receivables(USD)',
        'Uninvoiced Final Receivables(USD)',
        'Summary Outstanding Receivables(USD)',
        'Sales Contractual Counterparty Open Account Approved',
        'Purchase Dry Weight', 'Purchase Dry Weight UOM', 'Purchase Wet Weight',
        'Purchase Wet Weight UOM', 'Receipt In Progress',
        'Delivery In Progress', 'Eligible Dry Weight', 'Eligible Wet Weight']

    dff = dff.drop(columns = ['Sales Provisional Final2'])

    dff.columns = ['Purchase Group Company Code', 'Purchase Trade ID',
        'Purchase Quota Titan ID', 'Purchase Assignment Reference',
        'Sales Assignment Reference', 'Purchase Traffic Operator',
        'Purchase Contractual Counterparty', 'Purchase Commodity',
        'Sales Assignment Counterparty', 'Sales Credit Security Type',
        'LC Reference', 'LC Original Amount', 'LC Current Balance',
        'LC Issuance Date', 'LC Expiry Date', 'LC Opening Bank',
        'LC Confirming Bank', 'Total Financeable Sales Invoices Covered by LC',
        'LC Cover Available to SCR (SCR-Sales Contractual Rights)',
        'SCR Covered By LC (SCR-Sales Contractual Rights)', 'Sales Invoice ID',
        'Sales Provisional Final',
        'Sales Provisional Number', 'Sales Provisional Percentage',
        'Sales Invoice Date', 'Sales Invoice Due Date',
        'Sales Invoice Days Overdue', 'Sales Invoice Counterparty',
        'Cost Code ID', 'Cost Amount(USD)', 'Invoice Line Quantity',
        'Invoice Line Quantity UOM', 'Invoice Line Amount(USD)', 'Paid Amount',
        'Sales Assignment Eligible Percentage', 'Authorised Paid Quantity', 'Financed Quantity', 'Financing Ratio',
        'Unpaid Provisional Receivables(USD)', 'Unpaid Final Receivables(USD)',
        'Uninvoiced Final Receivables(USD)',
        'Summary Outstanding Receivables(USD)',
        'Sales Contractual Counterparty Open Account Approved',
        'Purchase Dry Weight', 'Purchase Dry Weight UOM', 'Purchase Wet Weight',
        'Purchase Wet Weight UOM', 'Receipt In Progress',
        'Delivery In Progress', 'Eligible Dry Weight', 'Eligible Wet Weight']

    cols_ffca_sales = ['Sales Assignment ID ', 'Estimated Sales Valuation']
    df_FFCA_s = pd.DataFrame(df_ffca, columns=cols_ffca_sales)
    df_FFCA_s['Sales Assignment ID '] = df_FFCA_s['Sales Assignment ID '].astype(str)
    df_final_s = pd.merge(dff, df_FFCA_s, left_on='Sales Assignment Reference', right_on='Sales Assignment ID ', how="left")
    df_final_s = df_final_s.drop(columns=['Cost Amount(USD)', 'Sales Assignment ID '])
    df_final_s.columns = ['Purchase Group Company Code', 'Purchase Trade ID',
        'Purchase Quota Titan ID', 'Purchase Assignment Reference',
        'Sales Assignment Reference', 'Purchase Traffic Operator',
        'Purchase Contractual Counterparty', 'Purchase Commodity',
        'Sales Assignment Counterparty', 'Sales Credit Security Type',
        'LC Reference', 'LC Original Amount', 'LC Current Balance',
        'LC Issuance Date', 'LC Expiry Date', 'LC Opening Bank',
        'LC Confirming Bank', 'Total Financeable Sales Invoices Covered by LC',
        'LC Cover Available to SCR (SCR-Sales Contractual Rights)',
        'SCR Covered By LC (SCR-Sales Contractual Rights)', 'Sales Invoice ID',
        'Sales Provisional Final', 'Sales Provisional Number',
        'Sales Provisional Percentage', 'Sales Invoice Date',
        'Sales Invoice Due Date', 'Sales Invoice Days Overdue',
        'Sales Invoice Counterparty', 'Cost Code ID',
        'Invoice Line Quantity', 'Invoice Line Quantity UOM',
        'Invoice Line Amount(USD)', 'Paid Amount',
        'Sales Assignment Eligible Percentage', 'Authorised Paid Quantity',
        'Financed Quantity', 'Financing Ratio',
        'Unpaid Provisional Receivables(USD)', 'Unpaid Final Receivables(USD)',
        'Uninvoiced Final Receivables(USD)',
        'Summary Outstanding Receivables(USD)',
        'Sales Contractual Counterparty Open Account Approved',
        'Purchase Dry Weight', 'Purchase Dry Weight UOM', 'Purchase Wet Weight',
        'Purchase Wet Weight UOM', 'Receipt In Progress',
        'Delivery In Progress', 'Eligible Dry Weight', 'Eligible Wet Weight',
        'Cost Amount(USD)']
    df_final_s = df_final_s[order4]
    df_final_s = df_final_s.drop_duplicates(['Purchase Assignment Reference'])
    # df_final_s['Uninvoiced Final Receivables(USD)'] = df_final_s['Cost Amount(USD)'] - df_final_s['Invoice Line Amount(USD)']
    # df_final_s['Unpaid Provisional Receivables(USD)'] = df_final_s['Unpaid Provisional Receivables(USD)_x']

    def function_a(a):
        if a < 0:
            return 0
        else:
            return a

    def function_Finqty(a,b):
        if a == 0 :
            return 0 
        else:
            return b

    def function_Finqty_result4(a):
        if a == 'inf' :
            return 0 
        else:
            return a

    
    df_final_s['Uninvoiced Final Receivables(USD)']  = df_final_s.apply(lambda x: function_a(x['Uninvoiced Final Receivables(USD)']), axis = 1)
    df_final_s['Financing Ratio']  = df_final_s.apply(lambda x: function_Finqty(x['Financed Quantity'], x['Financing Ratio']), axis = 1)


    # df_final_s['Summary Outstanding Receivables(USD)'] = df_final_s['Uninvoiced Final Receivables(USD)'] + df_final_s['Unpaid Final Receivables(USD)'] + df_final_s['Unpaid Provisional Receivables(USD)']

    final_result = pd.concat([df_final_s, result4], ignore_index=True)

    final_result = final_result.sort_values(by=['Sales Assignment Reference', 'Delivery In Progress'], ascending=[False, False], na_position= 'first') 
    # final_result = final_result[order]
    final_result = final_result[order4]
    final_result['Purchase Group Company Code'] = 'PTE'
    final_result[['Purchase Assignment Reference']] = final_result[['Purchase Assignment Reference']].astype('str')
    final_result['Purchase Trade ID'] = final_result['Purchase Assignment Reference'].map(lambda x: str(x)[:6])
    final_result.loc[:, 'Purchase Quota Titan ID'] = final_result['Purchase Assignment Reference'].apply(lambda st: st[st.rfind(".")-8: st.rfind(".")])

    final_result = pd.merge(final_result, df4_x, on='Purchase Assignment Reference')


    final_result['Financing Ratio'] = pd.to_numeric(final_result['Financing Ratio'], errors='coerce')
    final_result = final_result.round({'Financing Ratio': 2})
    final_result['Sales Provisional Final'] = np.nan
    final_result[['Financing Ratio']] = final_result[['Financing Ratio']].astype('str')
    final_result['Financing Ratio']  = final_result.apply(lambda x: function_Finqty_result4(x['Financing Ratio']), axis = 1)
    final_result['Sales Assignment Eligible Percentage'] = final_result['Sales Assignment Eligible Percentage'].apply(lambda x: format(x,".2%"))

    # final_result.to_excel(path + "\output.xlsx", sheet_name='sales receivable')
    writer = pd.ExcelWriter(path + '/final_test.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    final_result.to_excel(excel_writer=writer, sheet_name='sales receivables')
    writer.save()
    writer.close()


    #unpaid
    cols_input = ['Purchase Group Company Code', 'Purchase Assignment Reference', 'Purchase Settlement Valuation(USD)', 'Collateral Document Type']
    df_input = pd.DataFrame(input_unpaid, columns=cols_input)


    def HC(a):
        return a == "HC"
    def transport(a):
        return a != "HC"


    df_hc = df_input.loc[df_input['Collateral Document Type'].apply(HC)]
    df_transport = df_input.loc[df_input['Collateral Document Type'].apply(transport)]

    #storage

    cols_inventory = ['INVENTORY ID','COLLATERAL DOCUMENT DATE', 'PURCHASE ASSIGNMENT', 'DRY WEIGHT', 'WET WEIGHT', 'CITY']
    cols_wc = ['LOCATION_DESCRIPTION','Country' ,'RISK_CATEGORY']
    cols_inventory2 = ['TRANSPORT', 'INVENTORY ID', 'MATERIAL ORIGIN COUNTRIES', 'OPERATIONS OWNER', 'PRODUCT', 'LOCATION', 'PURCHASE COUNTERPARTY']

    df_inventory_f1 = pd.DataFrame(df_inventory, columns=cols_inventory)
    df_inventory2_f1 = pd.DataFrame(df_vessel, columns=cols_inventory2)
    df_WC_f1 = pd.DataFrame(df_WC, columns=cols_wc)
    df_WC_f1['RISK_CATEGORY'] = df_WC_f1['RISK_CATEGORY'].map(lambda x: str(x)[13:])


    inventory_combination = pd.merge(df_inventory_f1, df_inventory2_f1, on=['INVENTORY ID'])
    final_unpaid_WHS = pd.merge(inventory_combination,df_WC_f1, left_on = 'LOCATION', right_on= 'LOCATION_DESCRIPTION', how='outer')



    result_storage = pd.merge(df_hc, final_unpaid_WHS, left_on = 'Purchase Assignment Reference', right_on= 'PURCHASE ASSIGNMENT')

    result_storage['Collateral Reference'] = None
    result_storage[['Purchase Assignment Reference']] = result_storage[['Purchase Assignment Reference']].astype('str')
    result_storage['Purchase Trade ID'] = result_storage['Purchase Assignment Reference'].map(lambda x: str(x)[:6])
    result_storage.loc[:, 'Purchase Quota Titan ID'] = result_storage['Purchase Assignment Reference'].apply(lambda st: st[st.rfind(".")-8: st.rfind(".")])

    result_storage.loc[:, 'Eligible Dry Weight'] = result_storage['DRY WEIGHT']
    result_storage.loc[:, 'Eligible Wet Weight'] = result_storage['WET WEIGHT']
    result_storage.loc[:, 'Authorized Paid Quantity'] = result_storage['DRY WEIGHT']
    result_storage.loc[:, 'Financed Quantity'] = result_storage['DRY WEIGHT']
    result_storage.loc[:, 'Inventory Current Value(USD)'] = result_storage['Purchase Settlement Valuation(USD)']

    result_storage['Purchase Settlement Valuation(USD)'] = result_storage['Purchase Settlement Valuation(USD)'] * (-1)

    result_storage['Purchase Dry Weight UOM'] = 'MT'
    result_storage['Purchase Wet Weight UOM'] = 'MT'
    result_storage['Delivery In Progress'] = 'N'
    result_storage['Receipt In Progress'] = 'N'
    result_storage['Financed Ratio'] = '1'

    result_storage = result_storage.drop(columns = ['TRANSPORT', 'LOCATION_DESCRIPTION', 'PURCHASE ASSIGNMENT'])

    result_storage.columns = ['Purchase Group Company Code', 'Purchase Assignment Reference', 'Purchase Settlement Valuation(USD)', 'Collateral Document Type',
        'Purchase Inventory ID', 'Collateral Document Date', 'Purchase Dry Weight', 'Purchase Wet Weight',
        'Warehouse City', 'Purchase Material Origin Country', 'Purchase Traffic Operator', 'Purchase Commodity',
        'Warehouse Reference','Purchase Contractual Counterparty', 'Warehouse Country', 'Warehouse Category', 'Collateral Reference',
         'Purchase Trade ID',
        'Purchase Quota Titan ID', 'Eligible Dry Weight', 'Eligible Wet Weight',
        'Authorized Paid Quantity', 'Financed Quantity',
        'Inventory Current Value(USD)', 'Purchase Dry Weight UOM',
        'Purchase Wet Weight UOM', 'Delivery In Progress',
        'Receipt In Progress', 'Financing Ratio']

    order = ['Purchase Group Company Code', 'Purchase Trade ID', 'Purchase Quota Titan ID', 'Purchase Assignment Reference', 'Purchase Inventory ID', 'Purchase Traffic Operator', 'Purchase Contractual Counterparty',
            'Collateral Document Type', 'Collateral Reference', 'Collateral Document Date', 'Purchase Commodity', 'Purchase Material Origin Country', 'Warehouse Reference', 'Warehouse Category', 'Warehouse Country',
            'Warehouse City', 'Purchase Dry Weight', 'Purchase Dry Weight UOM', 'Purchase Wet Weight', 'Purchase Wet Weight UOM', 'Purchase Settlement Valuation(USD)', 'Receipt In Progress', 'Delivery In Progress',
            'Eligible Dry Weight', 'Eligible Wet Weight', 'Authorized Paid Quantity', 'Financed Quantity', 'Financing Ratio', 'Inventory Current Value(USD)' ]

    result_storage = result_storage[order]

    #transit
    cols_movement = ['ID', 'LOAD DATE','LOAD', 'DESTINATION', 'DESTINATION DATE']
    cols_inventory2 = ['TRANSPORT', 'INVENTORY ID', 'MATERIAL ORIGIN COUNTRIES', 'OPERATIONS OWNER', 'PRODUCT', 'PURCHASE COUNTERPARTY', 'MOVEMENT ID', 'MODE', 'RECEIPT INCOTERMS']

    df_movement_f1 = pd.DataFrame(df_movement, columns=cols_movement)
    df_movement_f1['LOAD'] = df_movement_f1['LOAD'].astype(str)
    df_inventory2_f1 = pd.DataFrame(df_vessel, columns=cols_inventory2)

    #deliver detail
    df_movement_f1.loc[:, 'Load Country'] = df_movement_f1['LOAD'].apply(lambda st: st[st.rfind("(")+1:st.rfind(")")])
    df_movement_f1.loc[:, 'Delivery Country'] = df_movement_f1['DESTINATION'].apply(lambda st: st[st.rfind("(")+1:st.find(")")])

    inventory_combination = pd.merge(df_inventory_f1, df_inventory2_f1, on=['INVENTORY ID'])
    final_unpaid_transit = pd.merge(inventory_combination, df_movement_f1, left_on = 'MOVEMENT ID', right_on = 'ID', how='outer')
    result_transport = pd.merge(df_transport, final_unpaid_transit, left_on = 'Purchase Assignment Reference', right_on= 'PURCHASE ASSIGNMENT')

    result_transport['Collateral Reference'] = None
    result_transport[['Purchase Assignment Reference']] = result_transport[['Purchase Assignment Reference']].astype('str')
    result_transport['Purchase Trade ID'] = result_transport['Purchase Assignment Reference'].map(lambda x: str(x)[:6])
    result_transport.loc[:, 'Purchase Quota Titan ID'] = result_transport['Purchase Assignment Reference'].apply(lambda st: st[st.rfind(".")-8: st.rfind(".")])

    result_transport.loc[:, 'Eligible Dry Weight'] = result_transport['DRY WEIGHT']
    result_transport.loc[:, 'Eligible Wet Weight'] = result_transport['WET WEIGHT']
    result_transport.loc[:, 'BL Date'] = result_transport['COLLATERAL DOCUMENT DATE']
    result_transport.loc[:, 'Authorized Paid Quantity'] = result_transport['DRY WEIGHT']
    result_transport.loc[:, 'Financed Quantity'] = result_transport['DRY WEIGHT']
    result_transport.loc[:, 'Inventory Current Value(USD)'] = result_transport['Purchase Settlement Valuation(USD)']

    result_transport['Purchase Settlement Valuation(USD)'] = result_transport['Purchase Settlement Valuation(USD)'] * (-1)

    result_transport['Purchase Dry Weight UOM'] = 'MT'
    result_transport['Purchase Wet Weight UOM'] = 'MT'
    result_transport['Delivery In Progress'] = 'N'
    result_transport['Receipt In Progress'] = 'N'
    result_transport['Financed Ratio'] = '1'

    result_transport = result_transport.drop(columns = ['PURCHASE ASSIGNMENT', 'ID', 'MOVEMENT ID', 'CITY'])

    result_transport.columns = ['Purchase Group Company Code', 'Purchase Assignment Reference', 'Purchase Settlement Valuation(USD)',
       'Collateral Document Type', 'Purchase Inventory ID', 'Collateral Document Date',
       'Purchase Dry Weight', 'Purchase Wet Weight', 'Vessel Name',
       'Purchase Material Origin Country', 'Purchase Traffic Operator', 'Purchase Commodity',
       'Purchase Contractual Counterparty', 'Purchase Transport Mode', 'Purchase Incoterm', 'Load  Date',
       'Load Location', 'Delivery Location', 'Expected Delivery Date', 'Load Country',
       'Delivery Country', 'Collateral Reference',
       'Purchase Trade ID',
       'Purchase Quota Titan ID', 'Eligible Dry Weight', 'Eligible Wet Weight','BL Date',
       'Authorized Paid Quantity', 'Financed Quantity',
       'Inventory Current Value(USD)', 'Purchase Dry Weight UOM',
       'Purchase Wet Weight UOM', 'Delivery In Progress',
       'Receipt In Progress', 'Financing Ratio']

    order = ['Purchase Group Company Code', 'Purchase Trade ID', 'Purchase Quota Titan ID', 'Purchase Assignment Reference', 'Purchase Inventory ID', 'Purchase Traffic Operator', 'Purchase Contractual Counterparty',
        'Collateral Document Type', 'Collateral Reference', 'Collateral Document Date', 'Purchase Commodity', 'Purchase Material Origin Country', 'Purchase Incoterm', 'Purchase Transport Mode', 'Vessel Name', 'Load  Date',
        'Load Location', 'Load Country','BL Date', 'Expected Delivery Date', 'Delivery Location', 'Delivery Country' ,'Purchase Dry Weight', 'Purchase Dry Weight UOM', 'Purchase Wet Weight', 'Purchase Wet Weight UOM', 'Purchase Settlement Valuation(USD)', 'Receipt In Progress', 'Delivery In Progress',
        'Eligible Dry Weight', 'Eligible Wet Weight', 'Authorized Paid Quantity', 'Financed Quantity', 'Financing Ratio', 'Inventory Current Value(USD)' ]

    result_transport = result_transport[order]

    writer = pd.ExcelWriter(path + '/final_test.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    result_storage.to_excel(excel_writer=writer, sheet_name='unpaid_storage')
    writer.save()
    writer.close()

    writer = pd.ExcelWriter(path + '/final_test.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    result_transport.to_excel(excel_writer=writer, sheet_name='unpaid_transit')
    writer.save()
    writer.close()

    # result_storage.to_excel(path + "\output.xlsx", sheet_name='unpaid_storage')
    # result_transport.to_excel(path + "\output.xlsx", sheet_name='unpaid_transport')


# setting button
btn_print = tk.Button(window, text='Report', command=reporting)
btn_print.place(x=50, y=50)

    

window.mainloop()